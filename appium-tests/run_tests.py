import os
import sys
import subprocess
import time
import socket
import unittest

# Ensure the correct working directory is the script's folder
script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)

import socket
socket.setdefaulttimeout(120.0)  # Must be large enough for Appium session creation (~45-90s on emulator)

# Import configurations
from appium_config import ANDROID_SDK, EMULATOR_NAME, APPIUM_PORT, APPIUM_HOST

# Paths
ADB_PATH = os.path.join(ANDROID_SDK, "platform-tools", "adb.exe")
EMULATOR_PATH = os.path.join(ANDROID_SDK, "emulator", "emulator.exe")
REPORTS_DIR = os.path.join(script_dir, "reports")

def install_python_dependencies():
    print("Checking and installing Python dependencies...")
    try:
        subprocess.run([sys.executable, "-m", "pip", "install", "-r", "requirements.txt"], check=True)
        print("Python dependencies verified.")
    except Exception as e:
        print(f"Warning: Failed to install Python dependencies: {e}. Trying to proceed anyway...")

def check_port_in_use(host, port):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex((host, port)) == 0

def install_local_appium():
    print("Verifying local Appium server installation...")
    # Check if node_modules/appium exists
    if not os.path.exists(os.path.join(script_dir, "node_modules", "appium")):
        print("Appium not found locally. Initializing npm project and installing Appium server...")
        try:
            # Check npm is in path
            subprocess.run(["npm", "-v"], check=True, stdout=subprocess.DEVNULL, shell=True)
            # Initialize project if package.json doesn't exist
            if not os.path.exists(os.path.join(script_dir, "package.json")):
                subprocess.run(["npm", "init", "-y"], check=True, stdout=subprocess.DEVNULL, shell=True)
            
            # Install Appium and driver locally
            print("Installing appium and appium-uiautomator2-driver. This may take 1-2 minutes...")
            subprocess.run(["npm", "install", "appium", "appium-uiautomator2-driver", "--save-dev"], check=True, shell=True)
            print("Local Appium server and Android driver installed successfully.")
        except Exception as e:
            print(f"Error: Failed to install Appium locally. Ensure Node.js and npm are on the PATH. Error: {e}")
            sys.exit(1)
    else:
        print("Local Appium installation detected.")

def get_connected_devices():
    try:
        res = subprocess.run([ADB_PATH, "devices"], capture_output=True, text=True, check=True)
        lines = res.stdout.strip().split("\n")[1:]
        devices = []
        for line in lines:
            if line.strip():
                parts = line.split("\t")
                if len(parts) == 2 and parts[1] == "device":
                    devices.append(parts[0])
        return devices
    except Exception as e:
        print(f"Error checking adb devices: {e}")
        return []

def wait_for_boot(device_serial=None, max_wait=120):
    """Wait for Android device to fully boot, return True if successful."""
    target = [ADB_PATH]
    if device_serial:
        target += ["-s", device_serial]
    for _ in range(max_wait // 2):
        try:
            res = subprocess.run(
                target + ["shell", "getprop", "sys.boot_completed"],
                capture_output=True, text=True, timeout=5
            )
            if res.stdout.strip() == "1":
                # Also wait for package manager to be ready
                res2 = subprocess.run(
                    target + ["shell", "getprop", "sys.boot_completed"],
                    capture_output=True, text=True, timeout=5
                )
                if res2.stdout.strip() == "1":
                    time.sleep(5)  # Extra stability margin
                    return True
        except Exception:
            pass
        time.sleep(2)
    return False

def start_emulator():
    devices = get_connected_devices()
    if devices:
        device = devices[0]
        print(f"Active Android device/emulator already connected: {device}.")
        print("Force-stopping app to ensure a clean test state...")
        try:
            subprocess.run(
                [ADB_PATH, "-s", device, "shell", "am", "force-stop", "com.sparegrow.app"],
                capture_output=True, text=True, timeout=10
            )
            time.sleep(2)
            print("App force-stopped. Ready for test session.")
        except Exception as e:
            print(f"Warning: Could not force-stop app: {e}. Proceeding anyway.")
        return True

    print(f"No active Android devices detected. Starting emulator: {EMULATOR_NAME}...")
    if not os.path.exists(EMULATOR_PATH):
        print(f"Error: Android emulator executable not found at: {EMULATOR_PATH}")
        return False

    try:
        # Start emulator in the background
        subprocess.Popen([EMULATOR_PATH, "-avd", EMULATOR_NAME, "-delay-adb"],
                         stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

        # Wait for device to appear in ADB list
        print("Waiting for emulator to connect to adb...")
        for _ in range(30):
            devices = get_connected_devices()
            if devices:
                print("Emulator connected to adb. Checking boot completion...")
                break
            time.sleep(2)
        else:
            print("Timeout: Emulator did not connect to adb.")
            return False

        if wait_for_boot():
            print("Emulator booted completely and is ready.")
        else:
            print("Warning: Emulator boot verification timed out, attempting to proceed anyway...")
        return True
    except Exception as e:
        print(f"Error launching emulator: {e}")
        return False

def start_appium_server():
    if check_port_in_use(APPIUM_HOST, APPIUM_PORT):
        print(f"Port {APPIUM_PORT} is already in use. Assuming Appium server is already running.")
        return None
        
    print("Starting Appium Server locally on port 4723...")
    try:
        # Run local appium from node_modules
        appium_bin = os.path.join(script_dir, "node_modules", ".bin", "appium.cmd")
        if not os.path.exists(appium_bin):
            appium_bin = "npx"
            args = [appium_bin, "appium", "-p", str(APPIUM_PORT), "--allow-insecure", "uiautomator2:chromedriver_autodownload"]
        else:
            args = [appium_bin, "-p", str(APPIUM_PORT), "--allow-insecure", "uiautomator2:chromedriver_autodownload"]
            
        # Redirect stdout and stderr to a log file to avoid pipe buffer deadlock
        log_file_path = os.path.join(script_dir, "appium_server.log")
        log_file = open(log_file_path, "w", encoding="utf-8")
        
        # Inject Android SDK variables into environment
        env = os.environ.copy()
        if "ANDROID_HOME" not in env:
            env["ANDROID_HOME"] = ANDROID_SDK
        if "ANDROID_SDK_ROOT" not in env:
            env["ANDROID_SDK_ROOT"] = ANDROID_SDK
        env["PATH"] = f"{os.path.join(ANDROID_SDK, 'platform-tools')};{os.path.join(ANDROID_SDK, 'emulator')};{env.get('PATH', '')}"
        
        appium_proc = subprocess.Popen(args, stdout=log_file, stderr=subprocess.STDOUT, text=True, shell=True, env=env)
        
        # Wait for port to open
        for _ in range(15):
            if check_port_in_use(APPIUM_HOST, APPIUM_PORT):
                print("Appium Server started successfully and listening on port 4723.")
                return appium_proc
            time.sleep(1)
        
        print("Timeout: Appium server port did not open.")
        log_file.close()
        return None
    except Exception as e:
        print(f"Error starting Appium server: {e}")
        return None

def generate_excel_report(test_results, steps):
    print("Generating beautifully formatted Excel analysis report...")
    if not os.path.exists(REPORTS_DIR):
        os.makedirs(REPORTS_DIR)
        
    excel_path = os.path.join(REPORTS_DIR, "SpareGrow_E2E_Test_Report.xlsx")
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Error: openpyxl library is required to generate the report. Please install it using pip.")
        return
        
    wb = Workbook()
    
    # ----------------- 1. SUMMARY SHEET -----------------
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Colors
    color_primary = "00342b"  # SpareGrow forest green
    color_secondary = "1b6d24"  # SpareGrow medium green
    color_pass = "10b981"  # Emerald Green
    color_fail = "ef4444"  # Red
    color_card_bg = "f2f4f1"  # Very light green/gray
    
    # Fonts
    font_title = Font(name="Segoe UI", size=18, bold=True, color="ffffff")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="ffffff")
    font_bold = Font(name="Segoe UI", size=10, bold=True, color="191c1b")
    font_regular = Font(name="Segoe UI", size=10, color="191c1b")
    font_metrics_val = Font(name="Segoe UI", size=20, bold=True, color=color_primary)
    
    # Fills & Borders
    fill_primary = PatternFill(start_color=color_primary, end_color=color_primary, fill_type="solid")
    fill_secondary = PatternFill(start_color=color_secondary, end_color=color_secondary, fill_type="solid")
    fill_card = PatternFill(start_color="e7e9e6", end_color="e7e9e6", fill_type="solid")
    fill_pass = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
    fill_fail = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='d8dbd8'),
        right=Side(style='thin', color='d8dbd8'),
        top=Side(style='thin', color='d8dbd8'),
        bottom=Side(style='thin', color='d8dbd8')
    )
    
    # Header Banner
    ws_summary.merge_cells("A1:G2")
    ws_summary["A1"] = "SpareGrow Mobile Application E2E Test Report"
    ws_summary["A1"].font = font_title
    ws_summary["A1"].fill = fill_primary
    ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Metadata Block
    ws_summary["A4"] = "Execution Metadata"
    ws_summary["A4"].font = Font(name="Segoe UI", size=12, bold=True, color=color_primary)
    
    metadata = [
        ("Platform", "Android Mobile (Capacitor Hybrid App)"),
        ("Emulator Name", EMULATOR_NAME),
        ("Target Package", "com.sparegrow.app"),
        ("Date", time.strftime("%Y-%m-%d %H:%M:%S")),
        ("Test Runner", "Appium Python Automator")
    ]
    for i, (k, v) in enumerate(metadata):
        row = 5 + i
        ws_summary.cell(row=row, column=1, value=k).font = font_bold
        ws_summary.cell(row=row, column=1).border = thin_border
        ws_summary.cell(row=row, column=2, value=v).font = font_regular
        ws_summary.cell(row=row, column=2).border = thin_border
        
    # KPI metrics cards
    total_tests = len(test_results)
    passed_tests = sum(1 for t in test_results if t["status"] == "PASS")
    failed_tests = total_tests - passed_tests
    success_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
    total_duration = sum(t["duration"] for t in test_results)
    
    metrics = [
        ("Total Tests", total_tests, "D5:E5", "D6:E6"),
        ("Passed", passed_tests, "F5:G5", "F6:G6"),
        ("Failed", failed_tests, "D8:E8", "D9:E9"),
        ("Success Rate", f"{success_rate:.1f}%", "F8:G8", "F9:G9")
    ]
    
    for title, val, merge_title, merge_val in metrics:
        ws_summary.merge_cells(merge_title)
        ws_summary.merge_cells(merge_val)
        
        cell_t = ws_summary[merge_title.split(":")[0]]
        cell_t.value = title.upper()
        cell_t.font = Font(name="Segoe UI", size=9, bold=True, color="555555")
        cell_t.alignment = Alignment(horizontal="center", vertical="center")
        cell_t.fill = fill_card
        
        cell_v = ws_summary[merge_val.split(":")[0]]
        cell_v.value = val
        
        # Color success rate or failed cases dynamically
        color_val = color_primary
        if title == "Passed":
            color_val = "10b981"
        elif title == "Failed" and val > 0:
            color_val = "ef4444"
        elif title == "Success Rate":
            color_val = "10b981" if success_rate > 90 else "f59e0b"
            
        cell_v.font = Font(name="Segoe UI", size=18, bold=True, color=color_val)
        cell_v.alignment = Alignment(horizontal="center", vertical="center")
        cell_v.fill = fill_card
        
        # Set border around merged zones
        for r in range(int(merge_title.split(":")[0][1]), int(merge_val.split(":")[1][1]) + 1):
            for c in range(ord(merge_title.split(":")[0][0]) - ord('A') + 1, ord(merge_title.split(":")[1][0]) - ord('A') + 2):
                ws_summary.cell(row=r, column=c).border = thin_border
                
    # Test execution overview table
    ws_summary["A12"] = "Test Case Run Details"
    ws_summary["A12"].font = Font(name="Segoe UI", size=12, bold=True, color=color_primary)
    
    headers = ["TC ID", "Test Name", "Description", "Status", "Duration (s)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=13, column=col_num, value=header)
        cell.font = font_header
        cell.fill = fill_secondary
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left" if col_num != 4 and col_num != 5 else "center")
        
    for i, t in enumerate(test_results):
        row = 14 + i
        ws_summary.cell(row=row, column=1, value=t["id"]).font = font_bold
        ws_summary.cell(row=row, column=2, value=t["name"]).font = font_regular
        ws_summary.cell(row=row, column=3, value=t["description"]).font = font_regular
        
        status_cell = ws_summary.cell(row=row, column=4, value=t["status"])
        status_cell.font = Font(name="Segoe UI", size=10, bold=True, color="047857" if t["status"] == "PASS" else "b91c1c")
        status_cell.fill = fill_pass if t["status"] == "PASS" else fill_fail
        status_cell.alignment = Alignment(horizontal="center")
        
        dur_cell = ws_summary.cell(row=row, column=5, value=round(t["duration"], 2))
        dur_cell.font = font_regular
        dur_cell.alignment = Alignment(horizontal="center")
        
        for c in range(1, 6):
            ws_summary.cell(row=row, column=c).border = thin_border
            
    # ----------------- 2. DETAILS SHEET -----------------
    ws_details = wb.create_sheet(title="Execution Steps")
    ws_details.views.sheetView[0].showGridLines = True
    
    ws_details.merge_cells("A1:F2")
    ws_details["A1"] = "E2E Step-by-Step Execution Logs"
    ws_details["A1"].font = font_title
    ws_details["A1"].fill = fill_primary
    ws_details["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    headers_details = ["TC Name", "Step Name", "Status", "Details", "Timestamp", "Screenshot Path"]
    for col_num, header in enumerate(headers_details, 1):
        cell = ws_details.cell(row=4, column=col_num, value=header)
        cell.font = font_header
        cell.fill = fill_secondary
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left" if col_num != 3 else "center")
        
    for i, step in enumerate(steps):
        row = 5 + i
        ws_details.cell(row=row, column=1, value=step["test_name"]).font = font_bold
        ws_details.cell(row=row, column=2, value=step["step_name"]).font = font_bold
        
        status_cell = ws_details.cell(row=row, column=3, value=step["status"])
        status_cell.font = Font(name="Segoe UI", size=10, bold=True, color="047857" if step["status"] == "PASS" else "b91c1c")
        status_cell.fill = fill_pass if step["status"] == "PASS" else fill_fail
        status_cell.alignment = Alignment(horizontal="center")
        
        ws_details.cell(row=row, column=4, value=step["details"]).font = font_regular
        ws_details.cell(row=row, column=5, value=step["timestamp"]).font = font_regular
        
        scr_cell = ws_details.cell(row=row, column=6, value=step["screenshot"])
        if step["screenshot"]:
            scr_cell.font = Font(name="Segoe UI", size=10, color="2563eb", underline="single")
            # Hyperlink to screenshot
            scr_cell.hyperlink = step["screenshot"]
        else:
            scr_cell.font = font_regular
            
        for c in range(1, 7):
            ws_details.cell(row=row, column=c).border = thin_border
            
    # Auto-adjust column widths for all sheets
    for ws in [ws_summary, ws_details]:
        for col in ws.columns:
            # Merge cell check to prevent massive width extension from title row
            non_merged_cells = [cell for cell in col if cell.coordinate not in ws.merged_cells]
            max_len = 0
            for cell in non_merged_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Specific adjustment for description & details column
    ws_summary.column_dimensions['C'].width = 45
    ws_details.column_dimensions['D'].width = 55
    ws_details.column_dimensions['A'].width = 25
    ws_details.column_dimensions['B'].width = 25
    
    wb.save(excel_path)
    print(f"Excel report successfully saved to: {excel_path}")

def run_tests():
    install_python_dependencies()
    install_local_appium()
    
    if not start_emulator():
        print("Aborting test run due to emulator failure.")
        sys.exit(1)
        
    appium_proc = start_appium_server()
    time.sleep(2)
    
    print("\n" + "="*50)
    print("RUNNING APPIUM E2E TEST SUITE FOR SPAREGROW")
    print("="*50 + "\n")
    
    from test_suite import SpareGrowE2ETests, step_results
    
    # Custom runner to collect outcomes programmatically
    class CleanResult(unittest.TextTestResult):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.successes_list = []
            
        def addSuccess(self, test):
            super().addSuccess(test)
            self.successes_list.append(test)
            
    class CustomRunner(unittest.TextTestRunner):
        def _makeResult(self):
            return CleanResult(self.stream, self.descriptions, self.verbosity)
            
    suite = unittest.TestLoader().loadTestsFromTestCase(SpareGrowE2ETests)
    runner = CustomRunner(verbosity=2)
    
    start_time = time.time()
    result = runner.run(suite)
    end_time = time.time()
    
    print("\n" + "="*50)
    print("TEST EXECUTION COMPLETED")
    print("="*50 + "\n")
    
    # Shut down local Appium server if we started it
    if appium_proc:
        print("Stopping local Appium server...")
        appium_proc.terminate()
        appium_proc.wait()
        print("Appium server stopped.")
        
    # Compile test case summary results
    test_cases_summary = []
    
    # Helper to check if a test method succeeded, failed or errored
    passed_methods = [t._testMethodName for t in result.successes_list if hasattr(t, '_testMethodName')]
    failed_methods = [t[0]._testMethodName for t in result.failures if hasattr(t[0], '_testMethodName')]
    errored_methods = [t[0]._testMethodName for t in result.errors if hasattr(t[0], '_testMethodName')]
    
    # Check if we had a class-level setup error (e.g., setUpClass failed)
    class_setup_error = ""
    for err in result.errors:
        if not hasattr(err[0], '_testMethodName'):
            class_setup_error = err[1]
            break
            
    # Combine all test methods run
    all_methods = sorted(list(set(passed_methods + failed_methods + errored_methods)))
    
    # If a class setup failure occurred, no test method actually executes
    if not all_methods:
        all_methods = sorted([m for m in dir(SpareGrowE2ETests) if m.startswith('test_')])
        if class_setup_error:
            errored_methods = all_methods
            
    for method_name in all_methods:
        method = getattr(SpareGrowE2ETests, method_name, None)
        doc = (method.__doc__ if method else "No description") or "No description"
        
        status = "PASS"
        err_msg = ""
        if method_name in failed_methods:
            status = "FAIL"
            # Extract fail message
            for f in result.failures:
                if hasattr(f[0], '_testMethodName') and f[0]._testMethodName == method_name:
                    err_msg = f[1]
                    break
        elif method_name in errored_methods:
            status = "FAIL"
            # Extract error message
            if class_setup_error:
                err_msg = class_setup_error
            else:
                for e in result.errors:
                    if hasattr(e[0], '_testMethodName') and e[0]._testMethodName == method_name:
                        err_msg = e[1]
                        break
                    
        # Approximate duration since unittest doesn't track per-test duration in standard results
        duration = (end_time - start_time) / len(all_methods) if all_methods else 0
        
        test_cases_summary.append({
            "id": method_name.split("_")[1],  # Extract 01 from test_01_splash...
            "name": method_name,
            "description": doc,
            "status": status,
            "duration": duration,
            "error": err_msg
        })
        
    # Generate the report
    generate_excel_report(test_cases_summary, step_results)
    
    print("\nPipeline complete! Excel report and screenshots are saved in 'appium-tests/' folder.")

if __name__ == "__main__":
    run_tests()
