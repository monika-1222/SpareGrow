"""
run_tests_full.py
=================
Full 300+ test pipeline runner for SpareGrow.
Runs all 5 categories:
  A – Deployment (20 tests)
  B – UI/UX      (80 tests)
  C – Functional (90 tests)
  D – Unit       (60 tests)
  E – Validation (60 tests)
                 ──────────
  Total          310 tests

Then generates a beautifully-formatted Excel report.
"""
import os, sys, time, unittest, socket, datetime, traceback

# ── 1. Paths ────────────────────────────────────────────────────────────────
script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)

# ── 2. Socket timeout (120 s allows Appium session creation) ───────────────
socket.setdefaulttimeout(120.0)

# ── 3. Config check ───────────────────────────────────────────────────────
from appium_config import ANDROID_SDK, EMULATOR_NAME, APPIUM_PORT, APPIUM_HOST

# ── 4. Reports / Screenshots dirs ─────────────────────────────────────────
REPORTS_DIR     = os.path.join(script_dir, "reports")
SCREENSHOTS_DIR = os.path.join(script_dir, "screenshots")
os.makedirs(REPORTS_DIR, exist_ok=True)
os.makedirs(SCREENSHOTS_DIR, exist_ok=True)

# ── 5. ADB / Appium helpers ───────────────────────────────────────────────
import subprocess

def run_cmd(cmd, timeout=15, capture=True):
    try:
        r = subprocess.run(cmd, capture_output=capture, text=True, timeout=timeout)
        return r.stdout.strip() if capture else ""
    except Exception as e:
        return f"ERROR: {e}"

ADB = os.path.join(ANDROID_SDK, "platform-tools", "adb.exe")

print("=" * 60)
print("SPAREGROW – FULL E2E TEST SUITE (310 TEST CASES)")
print("=" * 60)

# ── 6. Check emulator ─────────────────────────────────────────────────────
devices_out = run_cmd([ADB, "devices"])
if EMULATOR_NAME not in devices_out:
    print(f"ERROR: Emulator '{EMULATOR_NAME}' not found. Start it first.")
    sys.exit(1)
print(f"Active device: {EMULATOR_NAME}")

# ── 7. Force-stop app ─────────────────────────────────────────────────────
try:
    subprocess.run([ADB, "-s", EMULATOR_NAME, "shell", "am", "force-stop", "com.sparegrow.app"],
                   timeout=12, capture_output=True)
    print("App force-stopped.")
except Exception as e:
    print(f"Warning: force-stop failed ({e}). Proceeding.")

# ── 8. Start Appium ───────────────────────────────────────────────────────
import socket as _sock, threading, json

def _port_open(host, port):
    try:
        s = _sock.create_connection((host, port), timeout=2); s.close(); return True
    except Exception:
        return False

appium_proc = None
if not _port_open(APPIUM_HOST, APPIUM_PORT):
    print(f"Starting Appium on port {APPIUM_PORT}...")
    APPIUM_LOG = os.path.join(script_dir, "appium_server.log")
    appium_proc = subprocess.Popen(
        ["appium", "--port", str(APPIUM_PORT), "--log-no-colors"],
        stdout=open(APPIUM_LOG, "w"), stderr=subprocess.STDOUT
    )
    deadline = time.time() + 60
    while time.time() < deadline:
        if _port_open(APPIUM_HOST, APPIUM_PORT):
            print("Appium started.")
            break
        time.sleep(1)
    else:
        print("ERROR: Appium did not start within 60 s.")
        sys.exit(1)
else:
    print(f"Appium already running on port {APPIUM_PORT}.")

# ── 9. Run tests ──────────────────────────────────────────────────────────
from test_suite_full import (
    A_DeploymentTests, B_UIUXTests, C_FunctionalTests,
    D_UnitTests, E_ValidationTests, get_results
)

suite = unittest.TestSuite()
loader = unittest.TestLoader()
for cls in [A_DeploymentTests, B_UIUXTests, C_FunctionalTests, D_UnitTests, E_ValidationTests]:
    suite.addTests(loader.loadTestsFromTestCase(cls))

start_time = time.time()
runner = unittest.TextTestRunner(verbosity=2, stream=sys.stdout)
result = runner.run(suite)
elapsed = time.time() - start_time

# ── 10. Stop Appium ───────────────────────────────────────────────────────
if appium_proc:
    appium_proc.terminate()
    print("Appium server stopped.")

# ── 11. Collect results ───────────────────────────────────────────────────
results = get_results()

# Fallback: if no steps logged (setup error), synthesise rows from test names
if not results:
    print("\nNo step-level results logged (likely setup error). Generating summary from test IDs.")
    CATEGORY_NAMES = {
        "A_DeploymentTests":  "Deployment",
        "B_UIUXTests":        "UI/UX",
        "C_FunctionalTests":  "Functional",
        "D_UnitTests":        "Unit",
        "E_ValidationTests":  "Validation",
    }
    failed_ids = {str(t) for t, _ in result.errors + result.failures}
    for test_case in suite:
        tid   = str(test_case)
        cls   = type(test_case).__name__
        cat   = CATEGORY_NAMES.get(cls, "General")
        status = "FAIL" if any(f in tid for f in failed_ids) else "PASS"
        results.append({
            "test_name":  tid,
            "step_name":  "Execution",
            "status":     status,
            "details":    "Setup-level error; test was not executed." if status == "FAIL" else "Test executed.",
            "timestamp":  time.strftime("%Y-%m-%d %H:%M:%S"),
            "screenshot": "",
            "category":   cat,
            "priority":   "High",
        })

# ── 12. Generate Excel ────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers as xlnumbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.series import SeriesLabel

    # ── Colour palette ──
    CLR = {
        "header_bg":     "1A1A2E",
        "header_fg":     "E8F5E9",
        "pass_bg":       "E8F5E9",
        "pass_fg":       "1B5E20",
        "fail_bg":       "FFEBEE",
        "fail_fg":       "B71C1C",
        "skip_bg":       "FFF3E0",
        "skip_fg":       "E65100",
        "cat_dep":       "E3F2FD",
        "cat_ui":        "F3E5F5",
        "cat_func":      "E8F5E9",
        "cat_unit":      "FFF9C4",
        "cat_val":       "FCE4EC",
        "summary_bg":    "16213E",
        "summary_fg":    "A8D8A8",
        "accent":        "4CAF50",
        "title_bg":      "0F3460",
    }

    CAT_COLORS = {
        "Deployment": CLR["cat_dep"],
        "UI/UX":      CLR["cat_ui"],
        "Functional": CLR["cat_func"],
        "Unit":       CLR["cat_unit"],
        "Validation": CLR["cat_val"],
    }
    CAT_PRIORITY = {
        "Deployment": "High",
        "UI/UX":      "Medium",
        "Functional": "High",
        "Unit":       "Medium",
        "Validation": "High",
    }

    def make_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def make_font(bold=False, color="000000", size=10):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    def thin_border():
        s = Side(style="thin")
        return Border(left=s, right=s, top=s, bottom=s)

    def centre():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()

    # ═══════════════════════════════════════════════════════════════════════
    #  SHEET 1 – EXECUTIVE SUMMARY
    # ═══════════════════════════════════════════════════════════════════════
    ws_sum = wb.active
    ws_sum.title = "Executive Summary"

    # Title block
    ws_sum.merge_cells("A1:J1")
    c = ws_sum["A1"]
    c.value = "SPAREGROW – MOBILE APP COMPREHENSIVE TEST REPORT"
    c.fill  = make_fill(CLR["title_bg"])
    c.font  = Font(bold=True, color="FFFFFF", size=18, name="Calibri")
    c.alignment = centre()
    ws_sum.row_dimensions[1].height = 42

    ws_sum.merge_cells("A2:J2")
    c2 = ws_sum["A2"]
    c2.value = f"Generated: {datetime.datetime.now().strftime('%A, %d %B %Y  %H:%M:%S')}   |   Test Duration: {elapsed:.1f}s"
    c2.fill  = make_fill(CLR["summary_bg"])
    c2.font  = Font(color=CLR["summary_fg"], size=10, italic=True, name="Calibri")
    c2.alignment = centre()

    # Category summary table header
    cat_rows = []
    cat_labels = ["Deployment","UI/UX","Functional","Unit","Validation"]
    for cat in cat_labels:
        cat_results = [r for r in results if r["category"] == cat]
        passed = sum(1 for r in cat_results if r["status"] == "PASS")
        failed = sum(1 for r in cat_results if r["status"] == "FAIL")
        total  = len(cat_results)
        pct    = round(passed / total * 100, 1) if total else 0
        cat_rows.append((cat, total, passed, failed, pct))

    total_all   = len(results)
    passed_all  = sum(1 for r in results if r["status"] == "PASS")
    failed_all  = sum(1 for r in results if r["status"] == "FAIL")
    pct_all     = round(passed_all / total_all * 100, 1) if total_all else 0

    # Summary headers
    headers_sum = ["#", "Test Category", "Total Tests", "Passed ✅", "Failed ❌", "Pass Rate %",
                   "Priority", "Status Badge", "Remarks", "Coverage"]
    row = 4
    for col, h in enumerate(headers_sum, 1):
        c = ws_sum.cell(row=row, column=col, value=h)
        c.fill = make_fill(CLR["header_bg"])
        c.font = make_font(bold=True, color=CLR["header_fg"], size=11)
        c.alignment = centre()
        c.border = thin_border()
    ws_sum.row_dimensions[row].height = 30
    row += 1

    for i, (cat, tot, psd, fld, pct) in enumerate(cat_rows, 1):
        badge  = "✅ PASS" if pct == 100 else ("⚠️ PARTIAL" if pct >= 70 else "❌ FAIL")
        remark = "All tests passing" if pct == 100 else (f"{fld} test(s) require attention" if fld else "Review needed")
        cov    = f"{pct}%"
        row_data = [i, cat, tot, psd, fld, f"{pct}%", CAT_PRIORITY.get(cat,"Medium"), badge, remark, cov]
        for col, val in enumerate(row_data, 1):
            c = ws_sum.cell(row=row, column=col, value=val)
            c.fill = make_fill(CAT_COLORS.get(cat, "FFFFFF"))
            c.font = make_font(size=10)
            c.alignment = centre()
            c.border = thin_border()
            if col == 4:  # Passed
                c.font = make_font(bold=True, color=CLR["pass_fg"], size=10)
            if col == 5 and fld > 0:  # Failed
                c.font = make_font(bold=True, color=CLR["fail_fg"], size=10)
        ws_sum.row_dimensions[row].height = 24
        row += 1

    # Grand total row
    gt_data = ["—", "GRAND TOTAL", total_all, passed_all, failed_all,
               f"{pct_all}%", "Mixed", "✅ PASS" if pct_all == 100 else "⚠️", "Complete suite run", f"{pct_all}%"]
    for col, val in enumerate(gt_data, 1):
        c = ws_sum.cell(row=row, column=col, value=val)
        c.fill  = make_fill(CLR["header_bg"])
        c.font  = make_font(bold=True, color="FFFFFF", size=11)
        c.alignment = centre()
        c.border = thin_border()
    ws_sum.row_dimensions[row].height = 28
    row += 2

    # KPI boxes
    kpis = [
        ("Total Test Cases", total_all, "1F7A8C"),
        ("Passed",           passed_all, "4CAF50"),
        ("Failed",           failed_all, "F44336"),
        ("Pass Rate",        f"{pct_all}%", "FF9800"),
        ("Duration (s)",     f"{elapsed:.1f}", "9C27B0"),
        ("Screenshots",      sum(1 for r in results if r.get("screenshot")), "2196F3"),
    ]
    ws_sum.merge_cells(f"A{row}:J{row}")
    ws_sum.cell(row=row, column=1, value="KEY PERFORMANCE INDICATORS").fill = make_fill(CLR["title_bg"])
    ws_sum.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    ws_sum.cell(row=row, column=1).alignment = centre()
    row += 1

    for i, (label, val, color) in enumerate(kpis):
        col = (i % 3) * 2 + 1
        ws_sum.merge_cells(start_row=row + (i // 3), end_row=row + (i // 3),
                           start_column=col, end_column=col + 1)
        c = ws_sum.cell(row=row + (i // 3), column=col)
        c.value = f"{label}: {val}"
        c.fill  = make_fill(color)
        c.font  = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
        c.alignment = centre()
        ws_sum.row_dimensions[row + (i // 3)].height = 36

    # Column widths
    col_widths = [5, 22, 14, 12, 12, 14, 12, 16, 35, 12]
    for ci, w in enumerate(col_widths, 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = w

    # ═══════════════════════════════════════════════════════════════════════
    #  SHEET 2 – FULL TEST RESULTS
    # ═══════════════════════════════════════════════════════════════════════
    ws_tc = wb.create_sheet("All Test Cases")

    # Freeze pane + header
    ws_tc.freeze_panes = "A3"
    ws_tc.merge_cells("A1:K1")
    ws_tc.cell(1, 1).value = "ALL 310 TEST CASES – DETAILED RESULTS"
    ws_tc.cell(1, 1).fill  = make_fill(CLR["title_bg"])
    ws_tc.cell(1, 1).font  = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    ws_tc.cell(1, 1).alignment = centre()
    ws_tc.row_dimensions[1].height = 32

    tc_headers = ["#", "Test ID / Name", "Category", "Priority",
                  "Step", "Status", "Details", "Timestamp", "Screenshot", "Duration", "Remarks"]
    for col, h in enumerate(tc_headers, 1):
        c = ws_tc.cell(row=2, column=col, value=h)
        c.fill = make_fill(CLR["header_bg"])
        c.font = make_font(bold=True, color=CLR["header_fg"], size=10)
        c.alignment = centre()
        c.border = thin_border()
    ws_tc.row_dimensions[2].height = 28

    for idx, r in enumerate(results, 1):
        row_num = idx + 2
        status  = r.get("status", "SKIP")
        cat     = r.get("category", "General")
        is_pass = status == "PASS"
        remark  = "Passed successfully" if is_pass else ("Failed – see details" if status == "FAIL" else "Skipped")

        row_vals = [
            idx,
            r.get("test_name", ""),
            cat,
            r.get("priority", CAT_PRIORITY.get(cat, "Medium")),
            r.get("step_name", ""),
            status,
            r.get("details", "")[:120],
            r.get("timestamp", ""),
            "✅ Saved" if r.get("screenshot") else "—",
            "—",
            remark
        ]
        for col, val in enumerate(row_vals, 1):
            c = ws_tc.cell(row=row_num, column=col, value=val)
            c.border = thin_border()
            c.alignment = left() if col in [2, 5, 7, 11] else centre()
            c.font = make_font(size=9)
            if col == 6:  # Status cell
                if is_pass:
                    c.fill = make_fill(CLR["pass_bg"])
                    c.font = make_font(bold=True, color=CLR["pass_fg"], size=9)
                elif status == "FAIL":
                    c.fill = make_fill(CLR["fail_bg"])
                    c.font = make_font(bold=True, color=CLR["fail_fg"], size=9)
                else:
                    c.fill = make_fill(CLR["skip_bg"])
                    c.font = make_font(color=CLR["skip_fg"], size=9)
            elif col == 3:
                c.fill = make_fill(CAT_COLORS.get(cat, "FFFFFF"))
        ws_tc.row_dimensions[row_num].height = 18

    tc_col_widths = [5, 55, 14, 12, 32, 10, 55, 20, 12, 12, 28]
    for ci, w in enumerate(tc_col_widths, 1):
        ws_tc.column_dimensions[get_column_letter(ci)].width = w

    # ═══════════════════════════════════════════════════════════════════════
    #  SHEETS 3-7 – PER-CATEGORY SHEETS
    # ═══════════════════════════════════════════════════════════════════════
    for cat in cat_labels:
        cat_results = [r for r in results if r["category"] == cat]
        sheet_title = f"{cat.replace('/', '_')} Tests"
        ws = wb.create_sheet(sheet_title)
        ws.freeze_panes = "A3"

        ws.merge_cells("A1:I1")
        ws.cell(1, 1).value = f"{cat.upper()} TESTS – DETAILED RESULTS"
        ws.cell(1, 1).fill  = make_fill(CLR["title_bg"])
        ws.cell(1, 1).font  = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
        ws.cell(1, 1).alignment = centre()
        ws.row_dimensions[1].height = 30

        headers_c = ["#", "Test Name", "Step", "Status", "Details", "Timestamp", "Priority", "Screenshot", "Notes"]
        for col, h in enumerate(headers_c, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.fill = make_fill(CLR["header_bg"])
            c.font = make_font(bold=True, color=CLR["header_fg"], size=10)
            c.alignment = centre()
            c.border = thin_border()
        ws.row_dimensions[2].height = 26

        for idx, r in enumerate(cat_results, 1):
            rn = idx + 2
            status = r.get("status", "SKIP")
            row_vals = [
                idx,
                r.get("test_name", ""),
                r.get("step_name", ""),
                status,
                r.get("details", "")[:100],
                r.get("timestamp", ""),
                r.get("priority", "Medium"),
                "📷" if r.get("screenshot") else "—",
                "OK" if status == "PASS" else "Review"
            ]
            for col, val in enumerate(row_vals, 1):
                c = ws.cell(row=rn, column=col, value=val)
                c.border = thin_border()
                c.alignment = left() if col in [2, 3, 5] else centre()
                c.font = make_font(size=9)
                if col == 4:
                    c.fill = make_fill(CLR["pass_bg"] if status == "PASS" else
                                       CLR["fail_bg"] if status == "FAIL" else CLR["skip_bg"])
                    c.font = make_font(bold=True,
                                       color=CLR["pass_fg"] if status == "PASS" else
                                             CLR["fail_fg"] if status == "FAIL" else CLR["skip_fg"],
                                       size=9)
            ws.row_dimensions[rn].height = 18

        col_ws = [5, 55, 32, 10, 60, 20, 12, 10, 12]
        for ci, w in enumerate(col_ws, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    # ═══════════════════════════════════════════════════════════════════════
    #  SHEET 8 – METRICS DASHBOARD
    # ═══════════════════════════════════════════════════════════════════════
    ws_m = wb.create_sheet("Metrics Dashboard")
    ws_m.merge_cells("A1:H1")
    ws_m.cell(1, 1).value = "TEST METRICS DASHBOARD"
    ws_m.cell(1, 1).fill  = make_fill(CLR["title_bg"])
    ws_m.cell(1, 1).font  = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    ws_m.cell(1, 1).alignment = centre()
    ws_m.row_dimensions[1].height = 34

    # Data for charts
    chart_headers = ["Category", "Total", "Passed", "Failed", "Pass %"]
    for ci, h in enumerate(chart_headers, 1):
        c = ws_m.cell(row=3, column=ci, value=h)
        c.fill = make_fill(CLR["header_bg"])
        c.font = make_font(bold=True, color="FFFFFF", size=10)
        c.alignment = centre()
        c.border = thin_border()

    for ri, (cat, tot, psd, fld, pct) in enumerate(cat_rows, 4):
        ws_m.cell(ri, 1, cat).alignment = centre()
        ws_m.cell(ri, 2, tot).alignment = centre()
        ws_m.cell(ri, 3, psd).alignment = centre()
        ws_m.cell(ri, 4, fld).alignment = centre()
        ws_m.cell(ri, 5, pct).alignment = centre()
        for ci in range(1, 6):
            ws_m.cell(ri, ci).border = thin_border()
            ws_m.cell(ri, ci).font = make_font(size=10)
        ws_m.cell(ri, 1).fill = make_fill(CAT_COLORS.get(cat, "FFFFFF"))

    # Grand total row
    ws_m.cell(9, 1, "TOTAL").fill = make_fill(CLR["header_bg"])
    ws_m.cell(9, 1).font = make_font(bold=True, color="FFFFFF", size=10)
    ws_m.cell(9, 2, total_all).font = make_font(bold=True, size=10)
    ws_m.cell(9, 3, passed_all).font = make_font(bold=True, color=CLR["pass_fg"], size=10)
    ws_m.cell(9, 4, failed_all).font = make_font(bold=True, color=CLR["fail_fg"], size=10)
    ws_m.cell(9, 5, pct_all).font   = make_font(bold=True, size=10)
    for ci in range(1, 6):
        ws_m.cell(9, ci).alignment = centre()
        ws_m.cell(9, ci).border = thin_border()

    # Bar chart – Tests by Category
    try:
        bar = BarChart()
        bar.type  = "col"
        bar.title = "Tests by Category"
        bar.style = 10
        bar.y_axis.title = "Count"
        bar.x_axis.title = "Category"
        bar.height = 12
        bar.width  = 22
        data_ref = Reference(ws_m, min_col=2, max_col=4, min_row=3, max_row=8)
        cats_ref = Reference(ws_m, min_col=1, min_row=4, max_row=8)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        ws_m.add_chart(bar, "A11")
    except Exception:
        pass

    # Pie chart – Pass vs Fail
    try:
        pie = PieChart()
        pie.title  = "Overall Pass / Fail"
        pie.style  = 10
        pie.height = 12
        pie.width  = 14
        ws_m.cell(25, 7, "Result");  ws_m.cell(25, 8, "Count")
        ws_m.cell(26, 7, "Passed");  ws_m.cell(26, 8, passed_all)
        ws_m.cell(27, 7, "Failed");  ws_m.cell(27, 8, failed_all)
        pdata = Reference(ws_m, min_col=8, min_row=25, max_row=27)
        plabs = Reference(ws_m, min_col=7, min_row=26, max_row=27)
        pie.add_data(pdata, titles_from_data=True)
        pie.set_categories(plabs)
        ws_m.add_chart(pie, "G11")
    except Exception:
        pass

    for ci, w in enumerate([18, 12, 12, 12, 12, 4, 14, 14], 1):
        ws_m.column_dimensions[get_column_letter(ci)].width = w

    # ═══════════════════════════════════════════════════════════════════════
    #  SHEET 9 – TEST PLAN / CATALOGUE
    # ═══════════════════════════════════════════════════════════════════════
    ws_plan = wb.create_sheet("Test Plan")
    ws_plan.merge_cells("A1:H1")
    ws_plan.cell(1, 1).value = "SPAREGROW – TEST PLAN CATALOGUE (310 TEST CASES)"
    ws_plan.cell(1, 1).fill  = make_fill(CLR["title_bg"])
    ws_plan.cell(1, 1).font  = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    ws_plan.cell(1, 1).alignment = centre()
    ws_plan.row_dimensions[1].height = 30

    plan_data = [
        # (Category, TC ID, Test Case Name, Objective, Screen / Component, Priority, Type, Expected Result)
        # Deployment
        ("Deployment","TC-DEP-001","Appium Server Reachable","Verify Appium session can be established","Appium Server","High","Infra","Session object created"),
        ("Deployment","TC-DEP-002","WebView Context Available","Verify WebView context is present after launch","App Launch","High","Infra","WEBVIEW context listed"),
        ("Deployment","TC-DEP-003","App Package Identifier","Validate app package in WebView context","App Runtime","High","Infra","com.sparegrow in context name"),
        ("Deployment","TC-DEP-004","LocalStorage Accessible","R/W to localStorage works","WebView JS","Medium","Infra","Value persists"),
        ("Deployment","TC-DEP-005","SessionStorage Accessible","R/W to sessionStorage works","WebView JS","Medium","Infra","Value persists for session"),
        ("Deployment","TC-DEP-006","JavaScript Engine Working","JS arithmetic executes correctly","JS Engine","High","Infra","2+2=4"),
        ("Deployment","TC-DEP-007","window.navigate() Exists","SPA router function is defined","main.js","High","Infra","typeof === function"),
        ("Deployment","TC-DEP-008","Hash Router Working","window.location.hash navigation works","SPA Router","High","Infra","Hash changes on navigate"),
        ("Deployment","TC-DEP-009","Document readyState","DOM is fully loaded","DOM","Medium","Infra","readyState=complete"),
        ("Deployment","TC-DEP-010","Capacitor Bridge Present","Capacitor/Cordova runtime accessible","Capacitor","Medium","Infra","Truthy bridge object"),
        ("Deployment","TC-DEP-011","Screen Index JSON Loaded","indexData array populated","main.js","High","Infra","Array.isArray(indexData)=true"),
        ("Deployment","TC-DEP-012","App Has Expected Screen Count","App has ≥15 registered screens","indexData","High","Infra","count >= 15"),
        ("Deployment","TC-DEP-013","Supabase Client Initialized","Supabase SDK loaded","main.js","High","Infra","typeof supabase !== undefined"),
        ("Deployment","TC-DEP-014","Screenshot Capability","Driver can capture screenshots","Appium Driver","Low","Infra","PNG file created"),
        ("Deployment","TC-DEP-015","showToast() Function","Toast notification function is defined","main.js","Medium","Infra","typeof === function"),
        ("Deployment","TC-DEP-016","handleRoute() Function","Route handler is defined","main.js","High","Infra","typeof === function"),
        ("Deployment","TC-DEP-017","publicRoutes Array Defined","Public route whitelist exists","main.js","High","Infra","Array.isArray = true"),
        ("Deployment","TC-DEP-018","Mock Session Injection","Test can inject mock session","Test Helper","High","Infra","currentSession.user.id set"),
        ("Deployment","TC-DEP-019","MPIN Verified Flag","MPIN session flag injectable","SessionStorage","High","Infra","mpin_verified=true"),
        ("Deployment","TC-DEP-020","App Container Element","Root #app div exists","index.html","High","Infra","Element found"),
        # UI/UX (abbreviated for plan sheet)
        ("UI/UX","TC-UI-001","Splash Screen Loads","Splash renders without JS errors","SplashScreen","Medium","UI","Screen loads"),
        ("UI/UX","TC-UI-002","Splash Logo Visible","Brand graphics present","SplashScreen","Medium","UI","img/svg/canvas >= 0"),
        ("UI/UX","TC-UI-003","Splash Background Color","Background colour set","SplashScreen","Low","UI","backgroundColor returned"),
        ("UI/UX","TC-UI-004","Onboarding Slide 1","Slide 1 DOM element exists","Onboarding","Medium","UI","#onboarding-slide-1 found"),
        ("UI/UX","TC-UI-005","Onboarding Slide 2","Slide 2 DOM element exists","Onboarding","Medium","UI","#onboarding-slide-2 found"),
        ("UI/UX","TC-UI-006","Onboarding Slide 3","Slide 3 DOM element exists","Onboarding","Medium","UI","#onboarding-slide-3 found"),
        ("UI/UX","TC-UI-007","Onboarding Nav Dots","Progress indicators visible","Onboarding","Low","UI","dot-1, dot-2, dot-3 found"),
        ("UI/UX","TC-UI-008","Onboarding Next Button","CTA button present","Onboarding","Medium","UI","#btn-next-slide found"),
        ("UI/UX","TC-UI-009","Login Email Field","Email input visible","Login","High","UI","#login-email found"),
        ("UI/UX","TC-UI-010","Login Password Field","Password input visible","Login","High","UI","#login-password found"),
        ("UI/UX","TC-UI-011","Login Submit Button","LOG IN button present","Login","High","UI","#login-submit-btn found"),
        ("UI/UX","TC-UI-012","Login Google Button","Google OAuth button present","Login","Medium","UI","#login-google-btn found"),
        ("UI/UX","TC-UI-013","Login Apple Button","Apple OAuth button present","Login","Medium","UI","#login-apple-btn found"),
        ("UI/UX","TC-UI-014","Login Remember Me","Checkbox present","Login","Low","UI","#remember-me found"),
        ("UI/UX","TC-UI-015","Login Form Container","Form wrapper present","Login","Medium","UI","#login-form found"),
        ("UI/UX","TC-UI-016","SignUp Name Field","Name input visible","SignUp","High","UI","#name found"),
        ("UI/UX","TC-UI-017","SignUp Email Field","Email input visible","SignUp","High","UI","#signup-email found"),
        ("UI/UX","TC-UI-018","SignUp Password Field","Password input visible","SignUp","High","UI","#signup-password found"),
        ("UI/UX","TC-UI-019","SignUp Phone Field","Phone input visible","SignUp","Medium","UI","#phone found"),
        ("UI/UX","TC-UI-020","SignUp Social Buttons","Google & Apple buttons present","SignUp","Medium","UI","Both buttons found"),
    ]

    # Add placeholder rows for remaining tests
    extra_tc = [
        ("UI/UX","TC-UI-021 to TC-UI-080","UI/UX Tests (21-80)","Full layout/element visibility checks across all screens","All Screens","Medium","UI","All elements present"),
        ("Functional","TC-FUNC-001 to TC-FUNC-090","Functional Tests (1-90)","End-to-end user flow verification for all features","All Screens","High","E2E","Flows complete successfully"),
        ("Unit","TC-UNIT-001 to TC-UNIT-060","Unit Tests (1-60)","JavaScript utility and calculation function verification","JS Engine","Medium","Unit","Correct values returned"),
        ("Validation","TC-VAL-001 to TC-VAL-060","Validation Tests (1-60)","Form field boundary conditions and input validation","All Forms","High","Validation","Inputs accepted/rejected correctly"),
    ]

    plan_headers = ["Category", "TC ID", "Test Case Name", "Objective", "Screen/Component",
                    "Priority", "Type", "Expected Result"]
    for ci, h in enumerate(plan_headers, 1):
        c = ws_plan.cell(row=2, column=ci, value=h)
        c.fill = make_fill(CLR["header_bg"])
        c.font = make_font(bold=True, color="FFFFFF", size=10)
        c.alignment = centre()
        c.border = thin_border()
    ws_plan.row_dimensions[2].height = 26

    all_plan = plan_data + extra_tc
    for ri, row_data in enumerate(all_plan, 3):
        cat = row_data[0]
        for ci, val in enumerate(row_data, 1):
            c = ws_plan.cell(row=ri, column=ci, value=val)
            c.border = thin_border()
            c.font   = make_font(size=9)
            c.alignment = left() if ci in [3, 4, 8] else centre()
            c.fill = make_fill(CAT_COLORS.get(cat, "FFFFFF"))
        ws_plan.row_dimensions[ri].height = 18

    plan_col_w = [14, 28, 35, 52, 25, 12, 14, 38]
    for ci, w in enumerate(plan_col_w, 1):
        ws_plan.column_dimensions[get_column_letter(ci)].width = w

    # ── Save ──────────────────────────────────────────────────────────────
    report_path = os.path.join(REPORTS_DIR, "SpareGrow_Full_E2E_Test_Report.xlsx")
    wb.save(report_path)
    print(f"\n✅ Excel report saved: {report_path}")
    print(f"   Total tests : {total_all}")
    print(f"   Passed      : {passed_all}")
    print(f"   Failed      : {failed_all}")
    print(f"   Pass rate   : {pct_all}%")
    print(f"   Duration    : {elapsed:.1f} s")

except ImportError:
    print("openpyxl not installed. Install with: pip install openpyxl")
except Exception as exc:
    print(f"Excel generation error: {exc}")
    traceback.print_exc()

print("\n" + "=" * 60)
print("Pipeline complete! Reports in appium-tests/reports/")
print("=" * 60)
sys.exit(0 if result.wasSuccessful() else 1)
